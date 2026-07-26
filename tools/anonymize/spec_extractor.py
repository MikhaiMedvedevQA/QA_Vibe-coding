"""Извлечение структуры полей ЭФ из Excel-спецификации в JSON-датасет.

Отделяет слой данных (стабильная структура полей) от слоя генерации (скиллы,
чек-листы, автотесты). Анонимизатор при обработке Excel сам отдаёт рядом с
`<stem>_anon.md` файл `<stem>_anon.json` — канонический датасет параметров ЭФ,
который затем потребляют скиллы (см. memory/skills-excel-input-plan.md).

Детекция столбцов — по подстрокам канонических заголовков проекта (маркеры
перенесены из памяти агентов required-fields-validator / default-values-validator,
см. .claude/agent-memory/...). Логика статусов (required/default/hidden) —
там же. Значения анонимизируются через mapper (как в тексте).
"""

import json
from pathlib import Path

from openpyxl.utils import get_column_letter

from extractors import _norm, ExcelContent


# --------------------------------------------------------------------------
# Справочник маркеров детекции столбцов (нормализованные подстроки, lower + ё->е).
# Порядок ролей = приоритет: для столбца берётся первая совпавшая роль.
# --------------------------------------------------------------------------

ROLE_MARKERS: list[tuple[str, list[str]]] = [
    ("default_value", [
        "значение, указываемое при открытии страницы",
        "значение по умолчанию",
        "значение при открытии",
        "default value",
        "default",
    ]),
    ("field_name", [
        "группа полей / поле",
        "группа полей",
        "название поля",
        "наименование поля",
    ]),
    ("type", [
        "тип поля",
        "тип элемента",
        "тип",
    ]),
    ("availability", [
        "доступность для редактирования",
        "доступность",
    ]),
    ("required", [
        "обязательность заполнения",
        "обязательность",
        "обязательное поле",
        "обязательное",
        "required",
        "mandatory",
    ]),
    ("input_method", [
        "способ указания значения",
        "способ указания",
        "способ заполнения",
        "способ ввода",
        "input method",
    ]),
    ("behavior_notes", [
        "прочие особенности поведения",
        "особенности поведения",
        "прочие особенности",
        "особенности",
    ]),
    ("constraints", [
        "текст ограничений",
        "оповещения об ошибках",
        "текст ошибки",
        "сообщение об ошибке",
        "подсказка",
        "tooltip",
        "ограничения",
    ]),
]

# Маркеры детекции строки-заголовка (первая строка, содержащая любой — заголовки).
HEADER_DETECT_MARKERS = [
    "обязательность",
    "тип поля",
    "способ указания",
    "значение, указываемое",
    "доступность для редактирования",
    "значение по умолчанию",
]

# Маркер скрытости поля — сканируется по всей строке (нормализованно).
HIDDEN_MARKERS = ["поле скрыто на эф", "скрыто на эф", "поле скрыто"]

# Условные маркеры обязательности/дефолта.
_REQUIRED_COND_MARKERS = ["если", "при ", "услов", "то поле", "то оно", "то оно обязательно"]
_DEFAULT_COND_MARKERS = ["если", "при услов"]
_DEFAULT_EMPTY_TOKENS = {"—", "-", "–", "—", "нет", "не предусмотрено", "нет значения", "n/a", "- нет -", "-нет-"}

SCHEMA_VERSION = "1.0"


def _norm_ci(s) -> str:
    """Нормализованная строка для сопоставления маркеров: lower, ё->е, strip."""
    if s is None:
        return ""
    return _norm(str(s)).replace("ё", "е").replace("Ё", "Е").lower().strip()


def _anon_text(text, mapper) -> str:
    """Анонимизировать текст ячейки через mapper (повторяет логику writers.anonymize_text)."""
    if text is None:
        return ""
    from writers import anonymize_text
    return anonymize_text(str(text), mapper)


# --------------------------------------------------------------------------
# Классификация статусов по исходным (неанонимизированным) значениям.
# --------------------------------------------------------------------------

def _classify_required(raw) -> str:
    r = _norm_ci(raw)
    if not r:
        return "unknown"
    if r == "обязательное":
        return "required"
    if "обязательное" in r:
        if any(k in r for k in _REQUIRED_COND_MARKERS):
            return "conditional"
        return "required"  # «Обязательное поле» и подобные
    if r in ("нет", "необязательное", "не обязательно", "нет обязательности"):
        return "no"
    if r.startswith("нет") and len(r) <= 8:
        return "no"
    return "conditional"  # иное непустое — условная обязательность


def _classify_default(raw) -> str:
    r = _norm_ci(raw)
    if not r:
        return "no_default"
    if r in _DEFAULT_EMPTY_TOKENS or r.startswith("—") or r.startswith("–") or r == "-":
        return "empty_default"
    if any(k in r for k in _DEFAULT_COND_MARKERS):
        return "needs_manual_check"
    return "has_default"


def _classify_availability(raw) -> str:
    r = _norm_ci(raw)
    if not r:
        return "unknown"
    if "недоступ" in r or r == "нет":
        return "not_available"
    if "только для чтения" in r or "read-only" in r or "read only" in r or "только чтение" in r:
        return "read_only"
    if "доступно" in r:
        return "available"
    return "conditional"


def _is_auto_input(raw) -> bool:
    r = _norm_ci(raw)
    return any(k in r for k in ["автоматически", "присваивается системой", "автоматич"])


def _find_hidden(row_cells) -> dict | None:
    """Найти маркер скрытости в любой ячейке строки. Возвращает evidence или None."""
    for cell in row_cells:
        if cell.value is None:
            continue
        r = _norm_ci(cell.value)
        if any(m in r for m in HIDDEN_MARKERS):
            return {
                "column": cell.column_letter,
                "row": cell.row,
                "text": str(cell.value).strip(),
            }
    return None


# --------------------------------------------------------------------------
# Разбор листа.
# --------------------------------------------------------------------------

def _detect_header_row(ws, scan_limit: int = 15) -> int | None:
    """Индекс (1-based) строки-заголовка либо None."""
    for row in ws.iter_rows(min_row=1, max_row=min(scan_limit, ws.max_row or 1), values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            r = _norm_ci(cell.value)
            if any(m in r for m in HEADER_DETECT_MARKERS):
                return cell.row
    return None


def _detect_roles(header_row_cells) -> dict[str, int]:
    """role -> column index (1-based). Для каждого столбца — первая совпавшая роль."""
    role_to_col: dict[str, int] = {}
    for cell in header_row_cells:
        if cell.value is None:
            continue
        h = _norm_ci(cell.value)
        if not h:
            continue
        for role, markers in ROLE_MARKERS:
            if any(m in h for m in markers):
                if role not in role_to_col:
                    role_to_col[role] = cell.column
                break
    return role_to_col


def _raw(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _extract_sheet(ws, mapper, sheet_index: int) -> dict:
    sheet_name_anon = _anon_text(ws.title, mapper)
    header_row_idx = _detect_header_row(ws)
    sheet_dict: dict = {
        # Сырое имя листа НЕ храним — оно содержит название продукта/СК.
        # Привязка к листу — по sheet_index + sheet_name_anon.
        "sheet_name_anon": sheet_name_anon,
        "sheet_index": sheet_index,
        "header_row": header_row_idx,
        "headers": {},
        "blocks": [],
        "parameters": [],
        "ambiguous_cases": [],
        "notes": [],
    }
    if header_row_idx is None:
        sheet_dict["notes"].append("Строка заголовков не обнаружена — лист пропущен (нет структуры спецификации).")
        return sheet_dict

    header_cells = list(ws[header_row_idx])
    role_to_col = _detect_roles(header_cells)
    sheet_dict["headers"] = {
        role: (next((str(c.value).strip() for c in header_cells if c.column == col), ""))
        for role, col in role_to_col.items()
    }

    if "field_name" not in role_to_col:
        sheet_dict["notes"].append("Не найден столбец имени поля — структура не распознана, лист пропущен.")
        return sheet_dict

    last_col_letter = get_column_letter(max(role_to_col.values()))

    current_block: str | None = None
    data_roles = ["type", "availability", "required", "input_method",
                  "default_value", "behavior_notes", "constraints"]

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        cells_by_col = {c.column: c for c in row}
        field_name_raw = _raw(cells_by_col.get(role_to_col["field_name"]))
        # Маркер скрытости может быть в любой ячейке строки.
        hidden_ev = _find_hidden(row)

        # Блок-разделитель: имя поля заполнено, все остальные роли пусты.
        other_filled = any(_raw(cells_by_col.get(role_to_col[r])) for r in data_roles if r in role_to_col)
        if field_name_raw and not other_filled:
            current_block = _anon_text(field_name_raw, mapper)
            sheet_dict["blocks"].append({
                # Сырое имя блока НЕ храним (может содержать продукт/СК).
                "name_anon": current_block,
                "row": row[0].row,
            })
            continue

        # Параметр.
        if not field_name_raw and not other_filled:
            # Пустая строка — пропускаем.
            continue

        if not field_name_raw and other_filled:
            # Данные без имени поля — неоднозначность.
            sheet_dict["ambiguous_cases"].append({
                "row": row[0].row,
                "reason": "Строка с признаками параметра, но без имени поля.",
                "values": {
                    r: _raw(cells_by_col.get(role_to_col[r]))
                    for r in data_roles if r in role_to_col and _raw(cells_by_col.get(role_to_col[r]))
                },
            })
            continue

        # Сбор значений по ролям.
        type_raw = _raw(cells_by_col.get(role_to_col.get("type")))
        avail_raw = _raw(cells_by_col.get(role_to_col.get("availability")))
        req_raw = _raw(cells_by_col.get(role_to_col.get("required")))
        input_raw = _raw(cells_by_col.get(role_to_col.get("input_method")))
        def_raw = _raw(cells_by_col.get(role_to_col.get("default_value")))
        beh_raw = _raw(cells_by_col.get(role_to_col.get("behavior_notes")))
        cons_raw = _raw(cells_by_col.get(role_to_col.get("constraints")))

        req_status = _classify_required(req_raw)
        def_status = _classify_default(def_raw)
        avail_status = _classify_availability(avail_raw)

        # Поправка из памяти: скрытое обязательное поле -> conditional.
        notes: list[str] = []
        if hidden_ev and req_status == "required":
            req_status = "conditional"
            notes.append("Поле скрыто на ЭФ и помечено обязательным -> условная обязательность.")

        row_idx = row[0].row
        def_col = role_to_col.get("default_value")
        def_cell = f"{get_column_letter(def_col)}{row_idx}" if def_col else None

        param = {
            "id": f"sheet{sheet_index}-row{row_idx}",
            "name": field_name_raw,
            "name_anon": _anon_text(field_name_raw, mapper),
            "block": current_block,
            "type": _anon_text(type_raw, mapper) if type_raw else "",
            "availability": {
                "raw": _anon_text(avail_raw, mapper) if avail_raw else "",
                "interpreted": avail_status,
            } if avail_raw else None,
            "required": {
                "raw": _anon_text(req_raw, mapper) if req_raw else "",
                "status": req_status,
                "evidence": {"column": get_column_letter(role_to_col["required"]), "row": row_idx}
                if "required" in role_to_col else None,
            },
            "input_method": {
                "raw": _anon_text(input_raw, mapper) if input_raw else "",
                "auto": _is_auto_input(input_raw),
            } if input_raw else None,
            "default_value": {
                "raw": _anon_text(def_raw, mapper) if def_raw else "",
                "interpreted": _anon_text(def_raw, mapper) if def_raw else "",
                "status": def_status,
                "source_cell": def_cell,
            },
            "behavior_notes": _anon_text(beh_raw, mapper) if beh_raw else "",
            "constraints": _anon_text(cons_raw, mapper) if cons_raw else "",
            "ui": {
                "hidden": bool(hidden_ev),
                "hidden_evidence": hidden_ev,
            },
            "source": {
                "sheet": sheet_name_anon,
                "row": row_idx,
                "cell_range": f"A{row_idx}:{last_col_letter}{row_idx}",
            },
            "notes": notes,
        }
        sheet_dict["parameters"].append(param)

    return sheet_dict


def build_spec_json(content: ExcelContent, mapper) -> dict:
    """Построить JSON-датасет параметров ЭФ из Excel-спецификации."""
    sheets = []
    for i, ws in enumerate(content.workbook.worksheets):
        sheets.append(_extract_sheet(ws, mapper, i))

    total_params = sum(len(s["parameters"]) for s in sheets)
    total_required = sum(
        1 for s in sheets for p in s["parameters"] if p["required"]["status"] == "required"
    )
    total_conditional = sum(
        1 for s in sheets for p in s["parameters"] if p["required"]["status"] == "conditional"
    )
    total_with_default = sum(
        1 for s in sheets for p in s["parameters"] if p["default_value"]["status"] == "has_default"
    )
    total_without_default = sum(
        1 for s in sheets for p in s["parameters"]
        if p["default_value"]["status"] in ("no_default", "empty_default")
    )
    total_needs_manual = sum(
        1 for s in sheets for p in s["parameters"] if p["default_value"]["status"] == "needs_manual_check"
    )
    total_hidden = sum(1 for s in sheets for p in s["parameters"] if p["ui"]["hidden"])
    total_ambiguous = sum(len(s["ambiguous_cases"]) for s in sheets)

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            # Сырые путь/имя файла НЕ храним: абсолютный путь утекает имя
            # пользователя ОС, имя файла — название продукта/СК. Храним только
            # анонимизированное имя (привязка к источнику — по нему + sheet_index).
            "file_name_anon": _anon_text(content.source.name, mapper),
        },
        "agent": {"name": "anonymizer-spec-extractor", "version": "1.0"},
        "sheets": sheets,
        "summary": {
            "total_sheets": len(sheets),
            "total_parameters": total_params,
            "required": total_required,
            "conditional_required": total_conditional,
            "with_default": total_with_default,
            "without_default": total_without_default,
            "default_needs_manual_check": total_needs_manual,
            "hidden_fields": total_hidden,
            "ambiguous_cases": total_ambiguous,
        },
    }


def write_spec_json(content: ExcelContent, out_dir: Path, mapper, out_name: str | None = None) -> Path:
    """Сохранить `<stem>_anon.json` рядом с md. out_name — без расширения (напр. 'Кейсы_anon')."""
    stem = content.source.stem
    base = out_name or f"{stem}_anon"
    out_file = out_dir / f"{base}.json"
    dataset = build_spec_json(content, mapper)
    out_file.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_file